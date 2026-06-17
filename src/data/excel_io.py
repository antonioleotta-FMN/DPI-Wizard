"""Import/export Excel dei Capital Market Assumptions (M2).

Serializza un CMASet su un file .xlsx strutturato e lo rilegge. Usa openpyxl.

Struttura del workbook:
  Foglio "AssetClass"   : una riga per asset class, colonne = campi del modello.
  Foglio "Correlazioni" : matrice quadrata con etichette su prima riga e prima colonna.
  Foglio "Metadati"     : nome e versione del set.

L'import RICOSTRUISCE i modelli di dominio passando dalla loro validazione: un file
malformato solleva un errore esplicito invece di produrre dati silenziosamente errati
(requisito "Precisione").
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.domain.models import AssetClass, CMASet, MatriceCorrelazione

_COLONNE_AC = [
    "nome", "mu_nominale", "mu_reale", "sigma", "costo", "duration",
    "illiquidita", "valuta", "copertura_valutaria", "peso_min", "peso_max",
]


def esporta_cma_excel(cma: CMASet, percorso: str | Path) -> Path:
    """Scrive un CMASet su file Excel. Restituisce il percorso scritto."""
    percorso = Path(percorso)
    wb = Workbook()

    # --- Foglio AssetClass ---
    ws = wb.active
    ws.title = "AssetClass"
    ws.append(_COLONNE_AC)
    for ac in cma.asset_class:
        ws.append([getattr(ac, c) for c in _COLONNE_AC])

    # --- Foglio Correlazioni ---
    wc = wb.create_sheet("Correlazioni")
    etich = cma.correlazioni.etichette
    wc.append([""] + etich)
    for nome, riga in zip(etich, cma.correlazioni.valori):
        wc.append([nome] + list(riga))

    # --- Foglio Metadati ---
    wm = wb.create_sheet("Metadati")
    wm.append(["chiave", "valore"])
    wm.append(["nome", cma.nome])
    wm.append(["versione", cma.versione])

    wb.save(percorso)
    return percorso


def _parse_bool(v: object) -> bool:
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    return str(v).strip().lower() in {"true", "vero", "1", "si", "sì", "x"}


def importa_cma_excel(percorso: str | Path) -> CMASet:
    """Legge un CMASet da file Excel, ricostruendo e validando i modelli."""
    percorso = Path(percorso)
    if not percorso.exists():
        raise FileNotFoundError(f"File non trovato: {percorso}")
    wb = load_workbook(percorso, data_only=True)

    for foglio in ("AssetClass", "Correlazioni", "Metadati"):
        if foglio not in wb.sheetnames:
            raise ValueError(f"Foglio mancante nel file: '{foglio}'")

    # --- AssetClass ---
    ws = wb["AssetClass"]
    righe = list(ws.iter_rows(values_only=True))
    if not righe:
        raise ValueError("Foglio 'AssetClass' vuoto")
    intestazione = [str(c).strip() if c is not None else "" for c in righe[0]]
    if intestazione[: len(_COLONNE_AC)] != _COLONNE_AC:
        raise ValueError(
            f"Intestazione AssetClass inattesa. Attese colonne: {_COLONNE_AC}"
        )
    asset_class = []
    for r in righe[1:]:
        if r is None or all(c is None for c in r):
            continue
        d = dict(zip(_COLONNE_AC, r))
        asset_class.append(
            AssetClass(
                nome=str(d["nome"]),
                mu_nominale=float(d["mu_nominale"]),
                mu_reale=None if d["mu_reale"] is None else float(d["mu_reale"]),
                sigma=float(d["sigma"]),
                costo=float(d["costo"]) if d["costo"] is not None else 0.0,
                duration=None if d["duration"] is None else float(d["duration"]),
                illiquidita=_parse_bool(d["illiquidita"]),
                valuta=str(d["valuta"]) if d["valuta"] else "EUR",
                copertura_valutaria=float(d["copertura_valutaria"])
                if d["copertura_valutaria"] is not None
                else 0.0,
                peso_min=float(d["peso_min"]) if d["peso_min"] is not None else 0.0,
                peso_max=float(d["peso_max"]) if d["peso_max"] is not None else 1.0,
            )
        )
    if not asset_class:
        raise ValueError("Nessuna asset class trovata nel file")

    # --- Correlazioni ---
    wc = wb["Correlazioni"]
    righe_c = list(wc.iter_rows(values_only=True))
    if not righe_c:
        raise ValueError("Foglio 'Correlazioni' vuoto")
    etichette = [str(c) for c in righe_c[0][1:] if c is not None]
    valori = []
    for r in righe_c[1:]:
        if r is None or r[0] is None:
            continue
        valori.append([float(x) for x in r[1 : 1 + len(etichette)]])
    correlazioni = MatriceCorrelazione(etichette=etichette, valori=valori)

    # --- Metadati ---
    wm = wb["Metadati"]
    meta = {}
    for r in wm.iter_rows(min_row=2, values_only=True):
        if r and r[0] is not None:
            meta[str(r[0]).strip()] = r[1]
    nome = str(meta.get("nome", "CMA importato"))
    versione = str(meta.get("versione", "1.0"))

    return CMASet(
        nome=nome,
        versione=versione,
        asset_class=asset_class,
        correlazioni=correlazioni,
    )
