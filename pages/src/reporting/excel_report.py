"""Reporting: esportazione Excel strutturata per il DPI (M4).

Genera un workbook con piu' fogli: metadati di elaborazione, comparto, assunzioni
(CMA), matrice di correlazione, pesi delle proposte e metriche/risultati a confronto.
Include un audit trail minimo (data di generazione, set CMA, versione).

Distingue i risultati del MODELLO dal testo deliberativo: l'export contiene solo dati
e metadati, non formula giudizi.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook

from src.domain.models import CMASet, Comparto, ConfigSimulazione, Proposta
from src.services import calcola_metriche, esegui_simulazione, verifica


def genera_report_excel(
    percorso: str | Path,
    cma: CMASet,
    comparto: Comparto,
    proposte: dict[str, Proposta],
    config: ConfigSimulazione,
    includi_simulazione: bool = True,
) -> Path:
    """Genera il report Excel completo. Restituisce il percorso scritto."""
    percorso = Path(percorso)
    wb = Workbook()

    # --- Metadati / audit trail ---
    ws = wb.active
    ws.title = "Metadati"
    ws.append(["chiave", "valore"])
    ws.append(["Generato il", datetime.now(timezone.utc).isoformat(timespec="seconds")])
    ws.append(["Set CMA", cma.nome])
    ws.append(["Versione CMA", cma.versione])
    ws.append(["Comparto", comparto.nome])
    ws.append(["Orizzonte (anni)", comparto.orizzonte_anni])
    ws.append(["Obiettivo", f"{comparto.obiettivo_rendimento} ({comparto.tipo_obiettivo.value})"])
    ws.append(["Definizione shortfall", comparto.shortfall.definizione.value])
    ws.append(["Inflazione", config.inflazione])
    ws.append(["N. simulazioni", config.n_simulazioni])
    ws.append(["Seed", config.seed])
    ws.append(["Confidenza VaR", config.confidenza_var])
    ws.append(["Nota", "Risultati di modello. Non costituiscono testo deliberativo."])
    if "[DEMO]" in cma.nome:
        ws.append(["AVVERTENZA", "Dati DEMO: non ufficiali del Fondo."])

    # --- Assunzioni ---
    wa = wb.create_sheet("Assunzioni")
    wa.append([
        "Asset class", "Rend. nominale", "Rend. reale", "Volatilita", "Costo",
        "Duration", "Illiquida", "Valuta", "Cop. valutaria", "Peso min", "Peso max",
    ])
    for ac in cma.asset_class:
        wa.append([
            ac.nome, ac.mu_nominale, ac.mu_reale, ac.sigma, ac.costo, ac.duration,
            ac.illiquidita, ac.valuta, ac.copertura_valutaria, ac.peso_min, ac.peso_max,
        ])

    # --- Correlazioni ---
    wc = wb.create_sheet("Correlazioni")
    etich = cma.correlazioni.etichette
    wc.append([""] + etich)
    for nome, riga in zip(etich, cma.correlazioni.valori):
        wc.append([nome] + list(riga))

    # --- Pesi delle proposte ---
    wp = wb.create_sheet("Pesi")
    nomi = [ac.nome for ac in cma.asset_class]
    wp.append(["Asset class"] + list(proposte.keys()))
    for n in nomi:
        wp.append([n] + [proposte[p].pesi.get(n, 0.0) for p in proposte])

    # --- Metriche e risultati ---
    wm = wb.create_sheet("Risultati")
    intest = [
        "Proposta", "Rend. nominale", "Rend. netto", "Rend. reale", "Volatilita",
        "Quota illiquida", "Esp. valutaria", "Duration media", "Stato vincoli",
    ]
    if includi_simulazione:
        intest += ["P(shortfall)", "VaR", "ES mercato"]
    wm.append(intest)

    for nome, p in proposte.items():
        m = calcola_metriche(cma, p, config.inflazione, comparto.orizzonte_anni)
        _, stato = verifica(p, comparto, cma)
        riga = [
            nome, m.rendimento_nominale, m.rendimento_netto_costi, m.rendimento_reale,
            m.volatilita, m.quota_illiquida, m.esposizione_valutaria_non_coperta,
            m.duration_media, stato.value,
        ]
        if includi_simulazione:
            res = esegui_simulazione(cma, p, comparto, config)
            riga += [res.prob_shortfall, res.var, res.expected_shortfall]
        wm.append(riga)

    wb.save(percorso)
    return percorso
