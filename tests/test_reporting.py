"""Test del modulo di reporting Excel (M4)."""

import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.data.demo import cma_demo, comparto_demo  # noqa: E402
from src.domain.models import ConfigSimulazione, Proposta  # noqa: E402
from src.reporting import genera_report_excel  # noqa: E402


def _proposte_demo():
    nomi = [ac.nome for ac in cma_demo().asset_class]
    peso = 1.0 / len(nomi)
    return {
        "AAS vigente": Proposta(nome="AAS vigente", pesi={n: peso for n in nomi}),
        "Prudenziale": Proposta(nome="Prudenziale", pesi={
            "Liquidita": 0.10, "Govt EUR": 0.45, "Corporate": 0.20,
            "Equity DM": 0.15, "Equity EM": 0.05, "Real assets": 0.05,
            "Private markets": 0.00,
        }),
    }


def test_report_crea_file_con_fogli(tmp_path):
    cma, comparto = cma_demo(), comparto_demo()
    config = ConfigSimulazione(n_simulazioni=2_000, seed=1)
    percorso = tmp_path / "report.xlsx"
    genera_report_excel(percorso, cma, comparto, _proposte_demo(), config,
                        includi_simulazione=True)
    assert percorso.exists()
    wb = load_workbook(percorso)
    for foglio in ("Metadati", "Assunzioni", "Correlazioni", "Pesi", "Risultati"):
        assert foglio in wb.sheetnames


def test_report_senza_simulazione_piu_veloce(tmp_path):
    cma, comparto = cma_demo(), comparto_demo()
    config = ConfigSimulazione(n_simulazioni=2_000, seed=1)
    percorso = tmp_path / "report2.xlsx"
    genera_report_excel(percorso, cma, comparto, _proposte_demo(), config,
                        includi_simulazione=False)
    wb = load_workbook(percorso)
    intest = [c.value for c in wb["Risultati"][1]]
    assert "P(shortfall)" not in intest


def test_report_avvertenza_demo(tmp_path):
    cma, comparto = cma_demo(), comparto_demo()
    config = ConfigSimulazione(n_simulazioni=1_000, seed=1)
    percorso = tmp_path / "report3.xlsx"
    genera_report_excel(percorso, cma, comparto, _proposte_demo(), config,
                        includi_simulazione=False)
    wb = load_workbook(percorso)
    valori = [r[0].value for r in wb["Metadati"].iter_rows()]
    assert "AVVERTENZA" in valori
